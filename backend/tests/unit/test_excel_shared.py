"""
Unit tests: app.excel.shared — the reusable workbook layout/styling
helpers introduced for the Version 1.2 reporting infrastructure
(Attendance Reports is the reference implementation; Results/Fees/
Timetable/Users reports reuse these same helpers later).
"""

from io import BytesIO

from openpyxl import load_workbook

from app.excel.shared import build_report_workbook


def _load(xlsx_bytes: bytes):
    return load_workbook(BytesIO(xlsx_bytes))


def test_produces_valid_xlsx_bytes():
    xlsx_bytes = build_report_workbook(
        report_title="Test Report",
        subtitle="All Departments",
        columns=["Student", "Attendance %"],
        rows=[["Alice", 90.0], ["Bob", 75.0]],
    )

    assert isinstance(xlsx_bytes, bytes)
    assert xlsx_bytes.startswith(b"PK")  # .xlsx is a zip archive


def test_title_subtitle_and_header_row_are_written():
    xlsx_bytes = build_report_workbook(
        report_title="Attendance Report",
        subtitle="Department: All Departments",
        columns=["Student", "Attendance %"],
        rows=[["Alice", 90.0]],
    )
    sheet = _load(xlsx_bytes).active

    assert sheet.cell(row=1, column=1).value == "Attendance Report"
    assert sheet.cell(row=2, column=1).value == "Department: All Departments"
    assert sheet.cell(row=5, column=1).value == "Student"
    assert sheet.cell(row=5, column=2).value == "Attendance %"
    assert sheet.cell(row=6, column=1).value == "Alice"
    assert sheet.cell(row=6, column=2).value == 90.0


def test_empty_rows_still_produces_a_valid_workbook_with_header_only():
    xlsx_bytes = build_report_workbook(
        report_title="Attendance Report", subtitle="All Departments", columns=["Student", "Attendance %"], rows=[]
    )
    sheet = _load(xlsx_bytes).active

    assert sheet.cell(row=5, column=1).value == "Student"
    assert sheet.cell(row=6, column=1).value is None
