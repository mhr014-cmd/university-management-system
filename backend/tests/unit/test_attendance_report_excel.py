"""
Unit tests: app.excel.attendance_report.generate_attendance_report_excel —
the reporting-infrastructure vertical slice's reference Excel generator.
"""

import uuid
from io import BytesIO

from openpyxl import load_workbook

from app.excel.attendance_report import generate_attendance_report_excel
from app.schemas.attendance import AttendanceReportEntry

_ALL_SCOPE = {"department": "All Departments", "semester": "All Semesters", "student": "All Students"}


def test_produces_valid_xlsx_with_summary_rows():
    summary = [
        AttendanceReportEntry(student_id=uuid.uuid4(), student_name="Alice Islam", percentage=92.5),
        AttendanceReportEntry(student_id=uuid.uuid4(), student_name="Bob Rahman", percentage=78.0),
    ]

    xlsx_bytes = generate_attendance_report_excel(_ALL_SCOPE, summary)
    sheet = load_workbook(BytesIO(xlsx_bytes)).active

    assert xlsx_bytes.startswith(b"PK")
    assert sheet.cell(row=5, column=1).value == "Student"
    assert sheet.cell(row=5, column=2).value == "Attendance %"
    assert sheet.cell(row=6, column=1).value == "Alice Islam"
    assert sheet.cell(row=6, column=2).value == 92.5
    assert sheet.cell(row=7, column=1).value == "Bob Rahman"


def test_produces_valid_xlsx_with_empty_summary():
    xlsx_bytes = generate_attendance_report_excel(_ALL_SCOPE, [])

    assert xlsx_bytes.startswith(b"PK")
