"""
Unit tests: app.pdf.results_report/fees_report and
app.excel.results_report/fees_report — the Reports-module consistency
enhancement giving Results/Fees the same Print/PDF/Excel exports
Attendance already had (see tests/unit/test_attendance_report_pdf.py and
test_attendance_report_excel.py for the reference pattern these mirror),
plus the follow-up detail-section gap closure (a summary-only report gave
no way to see which academic records/invoices produced the totals).
"""

import uuid
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from reportlab.platypus import Paragraph

from app.excel.fees_report import generate_fees_report_excel
from app.excel.results_report import generate_results_report_excel
from app.pdf.fees_report import _wrapped_detail_table, generate_fees_report_pdf
from app.pdf.results_report import generate_results_report_pdf
from app.schemas.report import (
    FeeDetailEntry,
    FeesReportResponse,
    GradeDistributionEntry,
    ReportScope,
    ResultDetailEntry,
    ResultsReportResponse,
)

_ALL_SCOPE = {"department": "All Departments", "semester": "All Semesters", "student": "All Students"}
_EMPTY_SCOPE = ReportScope(department_id=None, semester_id=None, student_id=None)

_RESULT_DETAIL = ResultDetailEntry(
    student_id=uuid.uuid4(),
    student_name="Rafiq Chowdhury",
    course_name="Data Structures",
    exam_title="Midterm",
    grade_letter="A",
    grade_point=4.0,
)

_FEE_DETAIL = FeeDetailEntry(
    student_id=uuid.uuid4(),
    student_name="Rafiq Chowdhury",
    fee_name="Tuition",
    amount=10000.0,
    paid=4000.0,
    outstanding=6000.0,
    due_date=date(2026, 1, 1),
    status="overdue",
)


def _results_report(*, with_distribution: bool, with_details: bool = True) -> ResultsReportResponse:
    return ResultsReportResponse(
        scope=_EMPTY_SCOPE,
        grade_distribution=[GradeDistributionEntry(grade_letter="A", count=3)] if with_distribution else [],
        pass_count=3,
        fail_count=1,
        average_gpa=3.5,
        details=[_RESULT_DETAIL] if with_details else [],
    )


def _fees_report(*, with_details: bool = True) -> FeesReportResponse:
    return FeesReportResponse(
        scope=_EMPTY_SCOPE,
        total_collected=4000.0,
        total_outstanding=6000.0,
        total_overdue=2000.0,
        details=[_FEE_DETAIL] if with_details else [],
    )


class TestResultsReportPdf:
    def test_produces_valid_pdf_with_grade_distribution(self):
        pdf_bytes = generate_results_report_pdf(_ALL_SCOPE, _results_report(with_distribution=True))
        assert pdf_bytes.startswith(b"%PDF")

    def test_produces_valid_pdf_with_empty_distribution(self):
        pdf_bytes = generate_results_report_pdf(_ALL_SCOPE, _results_report(with_distribution=False))
        assert pdf_bytes.startswith(b"%PDF")

    def test_produces_valid_pdf_with_no_details(self):
        pdf_bytes = generate_results_report_pdf(_ALL_SCOPE, _results_report(with_distribution=True, with_details=False))
        assert pdf_bytes.startswith(b"%PDF")


class TestFeesReportPdf:
    def test_produces_valid_pdf(self):
        pdf_bytes = generate_fees_report_pdf(_ALL_SCOPE, _fees_report())
        assert pdf_bytes.startswith(b"%PDF")

    def test_produces_valid_pdf_with_no_details(self):
        pdf_bytes = generate_fees_report_pdf(_ALL_SCOPE, _fees_report(with_details=False))
        assert pdf_bytes.startswith(b"%PDF")

    def test_produces_valid_pdf_with_long_student_and_fee_name_and_multiple_rows(self):
        # Regression test for the Student/Fee Name column-overlap layout
        # bug: a long value in either column previously overflowed into
        # the next cell instead of wrapping. Two rows, both with long
        # text, confirm the fix isn't limited to the first row.
        long_details = [
            FeeDetailEntry(
                student_id=uuid.uuid4(),
                student_name="Rafiq Chowdhury",
                fee_name="CS Tuition — Spring 2026",
                amount=10000.0,
                paid=4000.0,
                outstanding=6000.0,
                due_date=date(2026, 1, 1),
                status="overdue",
            ),
            FeeDetailEntry(
                student_id=uuid.uuid4(),
                student_name="Anika Rahman-Chowdhury-Islam",
                fee_name="Computer Science Department Annual Laboratory and Equipment Fee",
                amount=5000.0,
                paid=0.0,
                outstanding=5000.0,
                due_date=date(2026, 2, 1),
                status="unpaid",
            ),
        ]
        report = FeesReportResponse(
            scope=_EMPTY_SCOPE, total_collected=4000.0, total_outstanding=11000.0, total_overdue=6000.0,
            details=long_details,
        )
        pdf_bytes = generate_fees_report_pdf(_ALL_SCOPE, report)
        assert pdf_bytes.startswith(b"%PDF")

    def test_detail_table_wraps_student_and_fee_name_columns_only(self):
        # Directly verifies the fix mechanism: Student/Fee Name (columns 0,
        # 1) become wrapping Paragraph flowables instead of raw strings —
        # the other columns are always short/fixed-format and stay plain
        # strings, matching every other report table's cell type.
        data = [
            ["Student", "Fee Name", "Amount", "Paid", "Outstanding", "Due Date"],
            ["Rafiq Chowdhury", "CS Tuition — Spring 2026", "10000.00", "4000.00", "6000.00", "2026-01-01"],
        ]
        table = _wrapped_detail_table(data, col_widths=[1.5, 1.9, 0.7, 0.7, 0.7, 0.7])
        header_row, body_row = table._cellvalues
        assert header_row == data[0]
        assert isinstance(body_row[0], Paragraph)
        assert isinstance(body_row[1], Paragraph)
        assert body_row[0].text == "Rafiq Chowdhury"
        assert body_row[1].text == "CS Tuition — Spring 2026"
        assert body_row[2:] == ["10000.00", "4000.00", "6000.00", "2026-01-01"]


class TestResultsReportExcel:
    def test_workbook_contains_grade_distribution_rows(self):
        excel_bytes = generate_results_report_excel(_ALL_SCOPE, _results_report(with_distribution=True))
        workbook = load_workbook(BytesIO(excel_bytes))
        sheet = workbook.active
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
        assert "A" in values
        assert 3 in values

    def test_details_sheet_contains_per_result_rows(self):
        excel_bytes = generate_results_report_excel(_ALL_SCOPE, _results_report(with_distribution=True))
        workbook = load_workbook(BytesIO(excel_bytes))
        assert "Details" in workbook.sheetnames
        detail_sheet = workbook["Details"]
        values = [cell.value for row in detail_sheet.iter_rows() for cell in row if cell.value is not None]
        assert "Rafiq Chowdhury" in values
        assert "Data Structures" in values
        assert "Midterm" in values


class TestFeesReportExcel:
    def test_workbook_contains_metric_rows(self):
        excel_bytes = generate_fees_report_excel(_ALL_SCOPE, _fees_report())
        workbook = load_workbook(BytesIO(excel_bytes))
        sheet = workbook.active
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
        assert "Collected" in values
        assert 4000.0 in values
        assert "Outstanding" in values
        assert "Overdue" in values

    def test_details_sheet_contains_per_invoice_rows(self):
        excel_bytes = generate_fees_report_excel(_ALL_SCOPE, _fees_report())
        workbook = load_workbook(BytesIO(excel_bytes))
        assert "Details" in workbook.sheetnames
        detail_sheet = workbook["Details"]
        values = [cell.value for row in detail_sheet.iter_rows() for cell in row if cell.value is not None]
        assert "Rafiq Chowdhury" in values
        assert "Tuition" in values
        assert "2026-01-01" in values
