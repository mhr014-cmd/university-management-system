"""
Shared Excel report scaffolding (reporting infrastructure).

Reusable workbook/worksheet layout and styling helpers for the new Reports
module — Attendance Reports (this vertical slice) is the reference
implementation; Results/Fees/Timetable/Users reports should reuse these
helpers rather than each hand-rolling openpyxl boilerplate, and rather
than forcing every report through one over-generic API.

openpyxl chosen as the Excel library: pure Python, no external binary
dependency, the de facto standard for server-side .xlsx generation —
same selection rationale already documented for reportlab in
requirements.txt.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_META_FONT = Font(color="6B7280", italic=True)

_TITLE_ROW = 1
_SUBTITLE_ROW = 2
_GENERATED_ROW = 3
_HEADER_ROW = 5


def autosize_columns(sheet: Worksheet, column_count: int, *, min_width: int = 12, max_width: int = 40) -> None:
    for col_index in range(1, column_count + 1):
        letter = get_column_letter(col_index)
        longest = max(
            (len(str(cell.value)) for cell in sheet[letter] if cell.value is not None),
            default=min_width,
        )
        sheet.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


def _write_merged_meta_row(sheet: Worksheet, row: int, value: str, column_count: int, font: Font) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(column_count, 1))
    cell = sheet.cell(row=row, column=1, value=value)
    cell.font = font


def write_header_row(sheet: Worksheet, columns: list[str], *, row: int = _HEADER_ROW) -> None:
    for col_index, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=col_index, value=column_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_data_sheet(sheet: Worksheet, *, title_row_value: str, columns: list[str], rows: list[list]) -> None:
    """Shared body for both the primary sheet (via build_report_workbook)
    and any extra_sheets (below) — a title/header row then data rows,
    auto-sized. Extracted so a detail sheet doesn't duplicate this loop."""
    _write_merged_meta_row(sheet, _TITLE_ROW, title_row_value, len(columns), _TITLE_FONT)
    write_header_row(sheet, columns, row=_TITLE_ROW + 1)

    for row_offset, row_values in enumerate(rows, start=1):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=_TITLE_ROW + 1 + row_offset, column=col_index, value=value)

    autosize_columns(sheet, len(columns))


def build_report_workbook(
    *,
    report_title: str,
    subtitle: str,
    columns: list[str],
    rows: list[list],
    # Optional additional worksheets — e.g. a "Details" sheet backing a
    # report's summary (Reports-module detail enhancement). Each tuple is
    # (sheet_title, columns, rows), reusing the same header/autosize
    # styling as the primary sheet. Every pre-existing caller (Attendance)
    # omits this, so its output is byte-for-byte unchanged.
    extra_sheets: list[tuple[str, list[str], list[list]]] | None = None,
) -> bytes:
    """Assembles a full .xlsx report: a title row, a subtitle row, a
    generated-on timestamp row, then a styled header row and the data
    rows the caller supplies. Column widths are auto-sized to content."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    _write_merged_meta_row(sheet, _TITLE_ROW, report_title, len(columns), _TITLE_FONT)
    _write_merged_meta_row(sheet, _SUBTITLE_ROW, subtitle, len(columns), _META_FONT)
    _write_merged_meta_row(
        sheet,
        _GENERATED_ROW,
        f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        len(columns),
        _META_FONT,
    )

    write_header_row(sheet, columns)

    for row_offset, row_values in enumerate(rows, start=1):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=_HEADER_ROW + row_offset, column=col_index, value=value)

    autosize_columns(sheet, len(columns))

    for sheet_title, extra_columns, extra_rows in extra_sheets or []:
        extra_sheet = workbook.create_sheet(title=sheet_title)
        _write_data_sheet(extra_sheet, title_row_value=sheet_title, columns=extra_columns, rows=extra_rows)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
